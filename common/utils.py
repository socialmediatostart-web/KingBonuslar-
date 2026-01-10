import datetime
import random
import shutil
import time
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from passlib.handlers.sha2_crypt import sha256_crypt
from common.constants import BASE_58_SYMBOLS, RequestReportTitles, RequestReportTotalTitles
import os
from pathlib import Path
import openpyxl
from common.google_drive import GoogleDriveWrapper


def get_base_58_string(length=20):
    return ''.join(random.choices(BASE_58_SYMBOLS, k=length))


def get_current_datetime():
    return datetime.datetime.now()


def get_datetime_after_seconds(seconds: int):
    return get_current_datetime() + datetime.timedelta(seconds=seconds)


def hash_password(password):
    return sha256_crypt.hash(password)


def verify_password(password, hashed):
    return sha256_crypt.verify(password, hashed)


class ReportGenerator:
    root_pwd = Path(os.getcwd(), 'root')
    report_filename = f"BonusRequests-{get_current_datetime().year}.xlsx"
    report_totals_tab_name = 'TOTALS'
    report_filepath = Path(root_pwd, report_filename)

    bold_font = Font(bold=True)
    gray_fill = PatternFill(
        fill_type="solid",
        start_color="DDDDDD",
        end_color="DDDDDD"
    )

    def is_running(self) -> bool:
        file_exists = self.root_pwd.exists()
        if not file_exists:
            os.mkdir(self.root_pwd)
        return file_exists

    def finish(self):
        shutil.rmtree(self.root_pwd)

    def upload_report(self):
        drive = GoogleDriveWrapper()
        root_files = drive.list_root_files()
        old_calendar_files = [f for f in root_files if f['name'] == self.report_filename]
        for file in old_calendar_files:
            drive.remove_file(file['id'])

        file_id = drive.upload_file(drive.root_folder_id, self.report_filepath)
        if file_id:
            return f"https://drive.google.com/file/d/{file_id}/view"

    def run_bonus_request_generation(self, bonus_requests_data: [], totals_user_data: [], top_referral_sources_data: {}):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove the default sheet

        # Organize data by month
        from collections import defaultdict
        from datetime import datetime

        data_by_month = defaultdict(list)
        current_year = get_current_datetime().year

        for record in bonus_requests_data:
            created_at = record.get(RequestReportTitles.request_created_at.value)
            if not isinstance(created_at, datetime):
                continue  # skip invalid dates
            if created_at.year != current_year:
                continue  # skip data not from current year
            month_key = created_at.strftime("%B")
            data_by_month[month_key].append(record)

        headers = [title.value for title in RequestReportTitles]

        for month, records in sorted(data_by_month.items()):
            ws = wb.create_sheet(title=month)
            ws.append(headers)
            ws.freeze_panes = "A2"

            # Header styling
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = self.bold_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = 22
                cell.fill = self.gray_fill

            # Add records
            for record in records:
                row = [
                    record.get(RequestReportTitles.request_created_at.value),
                    record.get(RequestReportTitles.user_created_at.value),
                    record.get(RequestReportTitles.request_status.value),
                    record.get(RequestReportTitles.tg_chat_id.value),
                    record.get(RequestReportTitles.site_id.value),
                    record.get(RequestReportTitles.group.value),
                    record.get(RequestReportTitles.is_subscribed.value),
                    record.get(RequestReportTitles.bonus_description.value),
                ]
                ws.append(row)

            # Center-align all data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add autofilter
            ws.auto_filter.ref = ws.dimensions
        # Start TOTALS tab creation
        wb.save(self.report_filepath)
        time.sleep(1)

        headers = [title.value for title in RequestReportTotalTitles]
        column_headers = headers[:-2]
        ref_sources_headers = headers[-2:]
        totals_ws = wb.create_sheet(self.report_totals_tab_name)
        for h in column_headers:
            totals_ws.append([h, totals_user_data.get(h, "")])

        totals_ws.append([''])
        totals_ws.append([h for h in ref_sources_headers])

        for row in top_referral_sources_data:
            totals_ws.append([row.get(RequestReportTotalTitles.top_referral_sources.value, ''),
                              row.get(RequestReportTotalTitles.referrals_count.value, '')])

        totals_ws.column_dimensions[get_column_letter(1)].width = 40
        totals_ws.column_dimensions[get_column_letter(2)].width = 22
        totals_ws.column_dimensions[get_column_letter(3)].width = 22

        for row_idx in range(1, len(column_headers)+1):
            cell = totals_ws.cell(row=row_idx, column=1)
            cell.font = self.bold_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

            cell.fill = self.gray_fill

        ref_sources_headers_row = len(column_headers)+2

        for col_idx in range(1, len(ref_sources_headers)+1):
            cell = totals_ws.cell(row=ref_sources_headers_row, column=col_idx)
            cell.font = self.bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell.fill = self.gray_fill

        wb.save(self.report_filepath)
        time.sleep(1)
